import json
import time
import openai
from xml.sax.saxutils import escape
import argparse
import openpyxl
import sys
import pandas as pd
import base64
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from openpyxl import Workbook

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.platypus import PageBreak

from dotenv import load_dotenv, find_dotenv
import os 
import logging

load_dotenv(find_dotenv())

OPEN_AI_API_KEY = os.getenv('OPEN_AI_API_KEY')
OPEN_AI_BASE_URL = os.getenv('OPEN_AI_BASE_URL')
OPEN_AI_API_TYPE = os.getenv('OPEN_AI_API_TYPE')
OPEN_AI_API_VERSION = os.getenv('OPEN_AI_API_VERSION')
OPEN_AI_DEPLOYMENT_ID = os.getenv('OPEN_AI_DEPLOYMENT_ID')
GMAIL_API_CREDENTIALS = os.getenv('GMAIL_API_CREDENTIALS')
COURSE_NAME = os.getenv('COURSE_NAME')
PROFESSOR_NAME = os.getenv('PROFESSOR_NAME')

deployment_id = OPEN_AI_DEPLOYMENT_ID

SCOPES = ['https://www.googleapis.com/auth/gmail.send']

logger = logging.getLogger()
logging.basicConfig(filename='student_processing.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
def setup_openai_api():
    openai.api_type = OPEN_AI_API_TYPE
    openai.api_base = OPEN_AI_BASE_URL
    openai.api_version = OPEN_AI_API_VERSION
    openai.api_key = OPEN_AI_API_KEY

system_config = {
    "role": "system",
    "content": f"You are an AI assistant of {COURSE_NAME} Course, you have to grade the student's homework and give them feedback. grade tem from 0 to 10."
}

user_config = {
    "role": "user",
    "content": "Context: '{content}'"
}

assistant_config = {
    "role": "assistant",
    "content": "Output:"
}

functions =[
   {
       "name": "create_json_payload",
         "type": "function",
            "description": "Create a json payload for the given input",
            
            "parameters": {
                "type": "object",
                "properties": {
                    "score": {
                        "type": "number",
                        "description": "The score for the student answer"
                    },
                    "feedback": {
                        "type": "string",
                        "description": "The feedback for the student answer"
                    }},
                "required": ["score", "feedback"]
                }   
   }
]
instruction = """Instructions:
    - You Ai assistant Computer Programming is here to help to assignment.
    - You evaluate Answer and give score and feedback.
    - Give score in number from 0 to 10.
    - Give feedback in string.
    """

def read_excel_file(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        logging.error(f"Error reading Excel file {file_path}: {e}")
        raise

def sanitize_text(text):
    return escape(str(text))

def chat_with_gpt(content):
    try:
        
        prompt = [
        {"role": "system", "content": instruction},
        {"role": "user", "content": f"""Context: {content}"""},
        {"role": "assistant", "content": "Output:"}]
        
        response = openai.ChatCompletion.create(deployment_id=deployment_id, 
                                            messages=prompt, 
                                            max_tokens=1024, 
                                            temperature=0.1, 
                                            functions=functions,
                                            function_call="auto")
        output = response['choices'][0]['message']['function_call']['arguments'].strip()
        return output
    except Exception as e:
        logger.error(f"Error in chat_with_gpt: {e}")
        raise 
def get_score_feedback(question, answer, wait_time=5):
    try:
        content = f"Question: {question} Answer: {answer}"
        output = chat_with_gpt(content)
        output = json.loads(output)
        time.sleep(wait_time)
        logger.info(f"Score and feedback generated for question {question} and answer {answer}")   
        return output['score'], output['feedback']
    except Exception as e:
        logging.error(f"Error getting score feedback for question {question} and answer {answer}: {e}")
        raise  
def generate_pdf(questions, answers, score_feedback, student_id, sum_score, total_questions, save_dir="./"):
    
    try:
        # Create a PDF document
        pdf_filename = f"{save_dir}/{student_id}_assignment.pdf"
        if os.path.exists(pdf_filename):
            version = 1
            while os.path.exists(pdf_filename):
                version += 1
                pdf_filename = f"{save_dir}/{student_id}_assignment_v{version}.pdf"


        doc = SimpleDocTemplate(pdf_filename, pagesize=letter)

        # Create a list to store the content of the PDF
        content = []

        # Define the styles for the text in the PDF
        styles = getSampleStyleSheet()
        
         # Add student ID and Total Score at the beginning
        content.append(Paragraph(f"Student ID: {student_id}", styles["Heading1"]))
        content.append(Paragraph(f"Total Score: {sum_score}/{total_questions*10}", styles["Heading1"]))
        content.append(Spacer(1, 24))  # Add some space after the student details


        # Iterate through the questions, answers, and feedback
        for i, (question, answer, (score, feedback)) in enumerate(zip(questions, answers, score_feedback), start=1):
            
            #make numpy.float64 to string
            answer = str(answer)
            # Add a question
            content.append(Paragraph(f"Question {i}:", styles["Heading1"]))
            content.append(Paragraph(question, styles["Normal"]))
            # Add an answer
            content.append(Spacer(1, 12))  # Add some space between question and answer
            content.append(Paragraph("Answer:", styles["Heading2"]))

            try:
                content.append(Paragraph(sanitize_text(answer), styles["Normal"]))
            except Exception as e:
                logging.error(f"Error sanitizing answer {answer}: {e}")
                content.append(Paragraph("Please you can see answers on googl form you submited ", styles["Normal"]))
            # Add the score and feedback
            content.append(Spacer(1, 12))  # Add some space between answer and score/feedback
            content.append(Paragraph("Score: {}/10".format(score), styles["Heading2"]))
            content.append(Paragraph("Feedback:", styles["Heading3"]))
            content.append(Paragraph(feedback, styles["Normal"]))

            # # Add a page break after each question
            # if i < len(questions):
            #     content.append(PageBreak())
        # Build the PDF document
        doc.build(content)
        return pdf_filename
    except Exception as e:
        logging.error(f"Error generating PDF for student_id {student_id}: {e}")
        raise

def send_pdf_email(email_address, pdf_path, service ):
    try:
    # Create the email
        message = MIMEMultipart()
        message['to'] = email_address
        message['subject'] = f'Your, Grades are Ready!'

        msg = MIMEText(f'Hi, This is Ai Assistant of your {PROFESSOR_NAME}. Your grades are ready. Please check the attached file. Thank you!')
        message.attach(msg)

        with open(pdf_path, 'rb') as f:
            attach = MIMEBase('application', 'octet-stream')
            attach.set_payload(f.read())

        encoders.encode_base64(attach)
        attach.add_header('Content-Disposition', f'attachment; filename= {pdf_path}')
        message.attach(attach)

        raw_msg = base64.urlsafe_b64encode(message.as_bytes()).decode('utf-8')
        body = {'raw': raw_msg}
        
        # Send the email
        service.users().messages().send(userId='me', body=body).execute()
    except Exception as e:
        logging.error(f"Error sending email to {email_address}: {e}")
        raise




def generate_excel_report(data, score_feedbacks):
    """
    Generate an Excel report containing student scores and feedback using score_feedbacks list.
    """
    workbook = Workbook()
    sheet = workbook.active

    # Create headers for the Excel sheet
    headers = ["Student Id", "Email", "Question", "Answer", "Score", "Feedback"]
    for idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=idx, value=header)

    row = 2  # Start from the second row for data

    # Extract questions from the dataframe
    questions = [question for question in data.columns if
                 question not in ["Timestamp", "Email Address", "Student Id ", "Score"]]

    for idx, student_data in data.iterrows():
        student_id = student_data["Student Id "]
        email = student_data["Email Address"]
        answers = student_data[questions].values

        for (question, answer), (score, feedback) in zip(zip(questions, answers), score_feedbacks[idx]):
            # Write the student data to the Excel sheet
            sheet.cell(row=row, column=1, value=student_id)
            sheet.cell(row=row, column=2, value=email)
            sheet.cell(row=row, column=3, value=question)
            sheet.cell(row=row, column=4, value=answer)
            sheet.cell(row=row, column=5, value=score)
            sheet.cell(row=row, column=6, value=feedback)

            row += 1  # Move to the next row for the next data entry

    # Save the workbook to a specified path
    report_path = "teacher_report.xlsx"
    workbook.save(report_path)
    return report_path

def generate_overall_report(data, score_feedbacks):
    """
    Generate an Excel report with overall scores, considering additional submissions.
    """
    workbook = Workbook()
    sheet = workbook.active
    
    # Create headers for the Excel sheet
    headers = ["Timestamp", "Student Id", "Email", "Questions Answered", "Total Marks", "Additional Marks (Later Submission)", "Total Score"]
    for idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=idx, value=header)
    
    row = 2  # Start from the second row for data

    student_ids = data["Student Id "].unique()

    for student_id in student_ids:
        student_submissions = data[data["Student Id "] == student_id].sort_values(by="Timestamp")
        total_marks = 0
        additional_marks = 0

        for _, student_data in student_submissions.iterrows():
            idx_feedback = data.index.get_loc(_)
            if idx_feedback < len(score_feedbacks):  # Check to ensure we don't go out of range
                student_feedback = score_feedbacks[idx_feedback]

                submission_score = sum([score for score, _ in student_feedback])
                if total_marks == 0:  # first submission
                    total_marks += submission_score
                else:  # later submissions
                    additional_score = submission_score * 0.5
                    additional_marks += additional_score
            else:
                print(f"No feedback found for student with ID: {student_id}")

        total_score = total_marks + additional_marks

        email = student_data["Email Address"]
        timestamp = student_data["Timestamp"]
        questions_answered = len([score for score, _ in score_feedbacks[data.index.get_loc(student_submissions.index[0])]])

        # Write data to the Excel sheet
        sheet.cell(row=row, column=1, value=timestamp)
        sheet.cell(row=row, column=2, value=student_id)
        sheet.cell(row=row, column=3, value=email)
        sheet.cell(row=row, column=4, value=questions_answered)
        sheet.cell(row=row, column=5, value=total_marks)
        sheet.cell(row=row, column=6, value=additional_marks)
        sheet.cell(row=row, column=7, value=total_score)
        
        row += 1  # Move to the next row for the next student
    
    # Save the workbook to a specified path
    report_path = "overall_report.xlsx"
    workbook.save(report_path)
    return report_path


def main(file_path, save_dir, wait_time):
    setup_openai_api()
    SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    credentials_json= GMAIL_API_CREDENTIALS   
    creds = None
    if not creds:
        flow = InstalledAppFlow.from_client_secrets_file(credentials_json, SCOPES)
        creds = flow.run_local_server(port=0)
    service = build('gmail', 'v1', credentials=creds)


    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    
    
    data = read_excel_file(file_path)
    questions = [ question for question in data.columns if question not in ["Timestamp", "Email Address", "Student Id ", "Score"]]
    answers_column = [column for column in data.columns if column not in ["Timestamp", "Email Address", "Student Id ", "Score"]]
    file_names = []
    email_addresses = data["Email Address"].values
    score_feedbacks = []
    student_processed = []
    for student_id, email_address in zip(data["Student Id "].values, email_addresses):
        try:
            logger.info(f"Processing student_id {student_id}")
            # check if the student has been processed before by studnet_id in current directory
            student_data = data[data["Student Id "] == student_id]
            answers = student_data[answers_column].values[0]
            score_feedback = [get_score_feedback(question, answer, wait_time) for question, answer in zip(questions, answers)]
            score_feedbacks.append(score_feedback)
            sum_score = sum([score for score, _ in score_feedback])

            total_questions = len(questions)
            pdf_filename = generate_pdf(questions, answers, score_feedback, student_id, sum_score, total_questions, save_dir)
            file_names.append(pdf_filename)
            logger.info(f"PDF generated for student_id {student_id}")
            send_pdf_email(email_address, pdf_filename, service)
            logger.info(f"Email sent to {email_address}")
            student_processed.append(student_id)
        except Exception as e:
            logging.error(f"Error sending email to {email_address}: {e}")
        time.sleep(3)
        
    # Generate the Excel report for the teacher
    try:
        generate_excel_report(data, score_feedbacks)
        logger.info("Excel report for the teacher generated successfully.")
    except Exception as e:
        logging.error(f"Error generating Excel report for the teacher: {e}")
        
    try:
        generate_overall_report(data, score_feedbacks)
        logger.info("Overall Excel report for the teacher generated successfully.")
    except Exception as e:
        logging.error(f"Error generating overall Excel report for the teacher: {e}")
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Grade student assignments using AI.')
    parser.add_argument('file_path', type=str, help='Path to the Excel file with student assignments')

      # New argument for directory to save generated PDFs
    parser.add_argument('--save_dir', type=str, default="./", help='Directory to save generated PDFs')

    # New argument for time to wait between processing each question
    parser.add_argument('--wait_time', type=float, default=0.0, help='Time (in seconds) to wait between processing each question')

    args = parser.parse_args()

    main(args.file_path, args.save_dir, args.wait_time)





